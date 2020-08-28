#!/usr/bin/env python
# coding: utf-8

# In[10]:


# Import necessary libraries
from openpyxl import Workbook
from openpyxl import load_workbook
from sklearn.metrics import cohen_kappa_score
import tkinter as tk
from tkinter import *
from tkinter.filedialog import askopenfile
import math


# In[13]:


# List of relevant values imported from an Excel file

file = askopenfile(mode="r", filetypes = [("Excel Files", "*.xlsx")])
# Insert the name of the XCEL list for the load_workbook
wb = load_workbook(file.name)
# Insert sheet name for the wb[" "] 
ws = wb["Sheet1"]
# Insert the 2 rater columns for the ws[" "] 
rater1 = ws["A"]
rater2 = ws["B"]
r1_val = [rater1[x].value for x in range(1,len(rater1))]
r2_val = [rater2[x].value for x in range(1,len(rater2))]
    
root = Tk()
root.destroy()
root.mainloop()


# In[18]:


# Determine if both values are non-zero and assigns 1

r1_n_val = []
r2_n_val = []

if len(r1_val) == sum(r1_val)+1:
    for x in r1_val:
        if x>0:
            r1_n_val.append(1)
        else: 
            r1_n_val.append(0)

    for x in r2_val:
        if x>0:
            r2_n_val.append(1)
        else: 
            r2_n_val.append(0)
    kappa_score = cohen_kappa_score(r1_n_val, r2_n_val)
    if math.isnan(kappa_score) == True:
        print(1.0)
    else:
        print(kappa_score)
else:
    kappa_score = cohen_kappa_score(r1_val, r2_val)
    if math.isnan(kappa_score) == True:
        print(1.0)
    else:
        print("The Cohen Kappa score is: "+str(kappa_score))
        
table = f"""
Poor           = < 0.00
Slight         = 0.00 - 0.20
Fair           = 0.20 - 0.40
Moderate       = 0.41 - 0.60
Substanstial   = 0.61 - 0.80
Almost Perfect = 0.81 - 1.00
"""
print(table)

